"""调试脚本：打印 Excel 解析后真正喂给大模型的内容（含合并单元格场景）。

用法：
    uv run python scripts/debug_excel_parse.py           # 内置样例：含合并单元格
    uv run python scripts/debug_excel_parse.py <路径>     # 解析你指定的真实 xlsx

关键点：extract_text_for_llm 返回的 text 就是 run_smart_import 里 build_messages
塞给 LLM 的原文（外层只包 「--- 文档内容开始/结束 ---」）。所以控制台打印的就是
大模型实际看到的输入。
"""

from __future__ import annotations

import sys
from pathlib import Path

from management_prd.services.file_text_extractor import extract_text_for_llm


def make_sample_with_merged_cells(path: Path) -> None:
    """构造含合并单元格 + 多种日期写法的样例。"""
    from datetime import datetime

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "需求"
    ws.append(["模块", "功能", "截止日期", "开始日期", "备注日期"])
    ws.append(["登录", "实现微信登录", datetime(2024, 11, 12), 45608.0, "2024年11月12日"])
    ws.append(["登录", "找回密码", datetime(2024, 11, 13), 45609.0, "2024/11/13"])
    ws.append(["支付", "微信支付", datetime(2024, 12, 1), 45627.0, "2024.12.01"])
    ws.append(["支付", "支付宝", datetime(2024, 12, 5), 45631.0, "2024-12-05"])
    # 列 D 故意设成日期格式，模拟「序列号 + 日期格式」即用户报告的 45608.0 场景
    for r in range(2, 6):
        ws.cell(row=r, column=4).number_format = 'yyyy"年"m"月"d"日"'
    # 纵向合并模块列，值保留在 A2
    ws.merge_cells("A2:A5")
    ws["A2"] = "登录与支付"
    wb.save(path)


def main() -> None:
    if len(sys.argv) > 1:
        target = Path(sys.argv[1])
    else:
        target = Path(__file__).with_name("_debug_merged_sample.xlsx")
        make_sample_with_merged_cells(target)
        print(f"# 已生成样例文件：{target}\n")

    text, fmt = extract_text_for_llm(target)
    print(f"=== source_format === {fmt}")
    print(f"=== 字符数 === {len(text)}")
    print("=== 喂给大模型的内容（extract_text_for_llm 返回值）===")
    print(text)


if __name__ == "__main__":
    main()