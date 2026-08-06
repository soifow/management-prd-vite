"""智能导入前置文件解析：把任意格式的需求文档转为 LLM 可读的纯文本。

二进制格式（.xlsx / .docx）用专用库解析为结构化 Markdown；其余回退纯文本读取。
仅在 :meth:`WebApi.pick_smart_import_file` 中调用，是 ``run_smart_import`` 之前的本地预处理。

返回 ``(text, source_format)`` 二元组：``text`` 为抽取后的纯文本，``source_format`` 为
格式标识（xlsx / docx / txt / md / csv ...），供前端提示「已识别为某格式」。

Excel 额外处理（避免污染 LLM 输入）：
- 合并单元格：被合并区单元格在转 Markdown 前填充左上角值，否则只剩左上角有值、
  其余为空，LLM 会丢失「该值对该合并区每行都生效」的语义。
- 日期归一：datetime 对象 / 日期序列号（如 45608）/ 中文日期串（``2024年11月12日``）
  统一转 ISO ``yyyy-MM-dd``，避免序列号原样喂给 LLM 无法理解。

设计详见 ``docs/design/smart-import-file-extraction.md``。
"""

from __future__ import annotations

import re
from collections.abc import Iterator
from datetime import date, datetime
from pathlib import Path
from typing import Any

from management_prd.errors import LlmError

# 纯文本直读的扩展名（含无扩展名）；其余在 _TEXT 未命中时同样回退纯文本读取。
_TEXT_EXTS = {".txt", ".md", ".markdown", ".csv", ".json", ".log", ".tsv", ""}


def extract_text_for_llm(path: Path) -> tuple[str, str]:
    """按扩展名把文件转为 ``(text, source_format)``；无法解析抛 LlmError。"""
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return _extract_xlsx(path)
    if ext == ".xls":
        return _extract_xls(path)
    if ext == ".docx":
        return _extract_docx(path)
    if ext == ".doc":
        raise LlmError("旧版 .doc 暂不支持，请用 Word 另存为 .docx 后重试")
    # 其余一律按文本读（含 .csv，LLM 原生理解；二进制会乱码，沿用现状）
    text = path.read_text(encoding="utf-8", errors="replace")
    return text, (ext.lstrip(".") or "txt")


# ──────────────────────────────────────────────────────────────
# Excel .xlsx
# ──────────────────────────────────────────────────────────────


def _extract_xlsx(path: Path) -> tuple[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise LlmError("Excel 解析依赖缺失，请联系开发者") from e
    try:
        # data_only=True 取缓存的计算值（需求文档多为直接录入文本）；data_only=False
        # 用于回退：缓存为 None 的公式单元格取公式串（决策 4）。
        # read_only=False：read_only 模式下 ws.merged_cells 不可用（探针实测返回空），
        # 为支持合并单元格填充，需求文档体积小，这里用全量加载。
        wb = load_workbook(path, data_only=True, read_only=False)
        wb_formula = load_workbook(path, data_only=False, read_only=False)
    except Exception as e:  # 损坏 / 加密
        raise LlmError(f"无法读取 Excel 文件：{e}") from e

    parts = [f"# 工作簿: {path.stem}（共 {len(wb.sheetnames)} 个工作表）\n"]
    for name in wb.sheetnames:
        ws = wb[name]
        ws_formula = wb_formula[name]
        parts.append(f"\n## 工作表: {name}\n")
        parts.append(_sheet_to_markdown(ws, ws_formula))
    wb.close()
    wb_formula.close()
    return "\n".join(parts).strip(), "xlsx"


def _sheet_to_markdown(ws: Any, ws_formula: Any) -> str:
    """把单个工作表渲染为 Markdown 表格；无有效行输出（空工作表）。

    处理顺序（决策见模块 docstring）：
    1. 合并单元格：被合并区单元格先填充左上角解析值（``_build_merge_fill_map``）。
    2. 日期归一：datetime / 日期序列号 / 中文日期串统一转 ISO ``yyyy-MM-dd``。
    3. 公式回退：值为空的单元格取公式 Workbook 的公式串（决策 4）。
    """
    fill_map = _build_merge_fill_map(ws, ws_formula)
    lines: list[str] = []
    for row in ws.iter_rows():
        cells: list[str] = []
        for cell in row:
            coord = (cell.row, cell.column)
            # 合并单元格：填充左上角值（已含日期归一 + 公式回退）
            if coord in fill_map:
                cells.append(_cell_str(fill_map[coord]))
                continue
            fcell = ws_formula.cell(row=cell.row, column=cell.column)
            value = _resolve_cell_value(cell.value, fcell.value, cell.number_format)
            cells.append(_cell_str(value))
        if not any(cells):
            continue  # 整行全空跳过
        lines.append("| " + " | ".join(cells) + " |")
    if not lines:
        return "（空工作表）"
    header = lines[0]
    col_count = header.count("|") - 1
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    return "\n".join([header, sep, *lines[1:]])


def _build_merge_fill_map(ws: Any, ws_formula: Any) -> dict[tuple[int, int], Any]:
    """构建合并单元格填充映射：(行, 列) -> 左上角单元格解析值。

    遍历 ``ws.merged_cells.ranges``，把每个合并区域除左上角外的所有坐标映射到
    左上角的解析值（含日期归一 + 公式回退）。渲染时被合并单元格直接取映射值，
    让 LLM 看到「该值对合并区域内每行都生效」的完整信息。
    """
    fill_map: dict[tuple[int, int], Any] = {}
    for rng in ws.merged_cells.ranges:
        top_data = ws.cell(row=rng.min_row, column=rng.min_col)
        top_formula = ws_formula.cell(row=rng.min_row, column=rng.min_col)
        top_value = _resolve_cell_value(
            top_data.value, top_formula.value, top_data.number_format
        )
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if r == rng.min_row and c == rng.min_col:
                    continue
                fill_map[(r, c)] = top_value
    return fill_map


def _resolve_cell_value(data_value: Any, formula_value: Any, number_format: Any) -> Any:
    """单元格取值：先日期归一，值空则回退公式串（决策 4）。"""
    if data_value is None:
        data_value = formula_value  # 公式缓存未算时回退公式串
    normalized = _normalize_date(data_value, number_format)
    if normalized is not None:
        return normalized
    return data_value


def _normalize_date(value: Any, number_format: Any) -> str | None:
    """把 Excel 日期统一转 ISO ``yyyy-MM-dd``；非日期返回 None。

    覆盖三种存储形态：
    1. datetime / date 对象（日期格式单元格，openpyxl 自动反序列化）。
    2. 日期序列号（如 45608）+ 日期/中文日期格式：经 ``from_excel`` 转日期。
    3. 纯文本日期串（``2024年11月12日`` / ``2024/11/12`` / ``2024.11.12`` 等）。
    """
    if isinstance(value, datetime):
        return value.date().isoformat()  # 需求日期不带时分，丢弃时间分量
    if isinstance(value, date):
        return value.isoformat()

    # 数值 + 日期格式：按 Excel 1900 序列号转日期（覆盖「45608.0 显示为中文日期」场景）
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if _is_date_format(str(number_format)):
            try:
                from openpyxl.utils.datetime import from_excel

                return from_excel(value).date().isoformat()
            except Exception:  # 超出合法日期序列范围等
                return None
        return None

    # 文本日期串：仅当整串就是一个日期才转（避免误伤描述性文本）
    if isinstance(value, str):
        return _parse_human_date(value)
    return None


def _is_date_format(number_format: str) -> bool:
    """判断 Excel 单元格数字格式是否为日期格式。

    先用 openpyxl 自带 ``is_date_format``（识别 y/m/d/h/s 等 token）；
    再补一道中文日期格式识别（``年/月/日``），覆盖 openpyxl 正则可能漏掉的自定义格式。
    """
    if not number_format or number_format == "General":
        return False
    try:
        from openpyxl.styles.numbers import is_date_format

        if is_date_format(number_format):
            return True
    except Exception:
        pass
    return "年" in number_format and "月" in number_format


# 文本日期串匹配：year(month)(day) 支持中文 / 斜杠 / 横线 / 点分隔；fullmatch 避免误伤
_CN_DATE = re.compile(r"(\d{4})年\s*(\d{1,2})月(?:\s*(\d{1,2})日)?")
_SLASH_DATE = re.compile(r"(\d{4})/(\d{1,2})/(\d{1,2})")
_DASH_DATE = re.compile(r"(\d{4})-(\d{1,2})-(\d{1,2})")
_DOT_DATE = re.compile(r"(\d{4})\.(\d{1,2})\.(\d{1,2})")


def _parse_human_date(value: str) -> str | None:
    """把纯文本日期串转 ISO ``yyyy-MM-dd``；非整串日期返回 None。"""
    s = value.strip()
    for pat in (_CN_DATE, _SLASH_DATE, _DASH_DATE, _DOT_DATE):
        m = pat.fullmatch(s)
        if not m:
            continue
        year, month = int(m.group(1)), int(m.group(2))
        day = int(m.group(3)) if m.lastindex and m.group(3) else 1
        try:
            return datetime(year, month, day).date().isoformat()
        except ValueError:
            continue
    return None


def _cell_str(value: object) -> str:
    """单元格渲染：None -> 空串；其余 str()。"""
    if value is None:
        return ""
    return str(value)


# ──────────────────────────────────────────────────────────────
# Excel .xls（旧版二进制，xlrd；2.x 仅支持 .xls）
# ──────────────────────────────────────────────────────────────


def _extract_xls(path: Path) -> tuple[str, str]:
    try:
        import xlrd
    except ImportError as e:
        raise LlmError("Excel(.xls) 解析依赖缺失，请联系开发者") from e
    try:
        book = xlrd.open_workbook(str(path))
    except Exception as e:  # 损坏 / 加密 / 新版 xlsx 误传为 .xls
        raise LlmError(f"无法读取 Excel 文件：{e}") from e

    parts = [f"# 工作簿: {path.stem}（共 {book.nsheets} 个工作表）\n"]
    for sheet in book.sheets():
        parts.append(f"\n## 工作表: {sheet.name}\n")
        parts.append(_xls_sheet_to_markdown(sheet, book.datemode))
    return "\n".join(parts).strip(), "xls"


def _xls_sheet_to_markdown(sheet: Any, datemode: int) -> str:
    """把 xlrd 工作表渲染为 Markdown 表格（首行作表头）；无有效行输出（空工作表）。

    与 :func:`_sheet_to_markdown` 同构：合并单元格填充（``sheet.merged_cells``，
    半开区间、0 基坐标）+ 日期归一（``xlrd.xldate``）。xlrd 无公式回退，直接取缓存值。
    """
    fill_map = _build_xls_merge_fill_map(sheet, datemode)
    lines: list[str] = []
    for ri in range(sheet.nrows):
        cells: list[str] = []
        for ci in range(sheet.ncols):
            if (ri, ci) in fill_map:
                cells.append(_cell_str(fill_map[(ri, ci)]))
                continue
            value = _normalize_xls_value(
                sheet.cell_value(ri, ci), sheet.cell_type(ri, ci), datemode
            )
            cells.append(_cell_str(value))
        if not any(cells):
            continue  # 整行全空跳过
        lines.append("| " + " | ".join(cells) + " |")
    if not lines:
        return "（空工作表）"
    header = lines[0]
    col_count = header.count("|") - 1
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    return "\n".join([header, sep, *lines[1:]])


def _build_xls_merge_fill_map(sheet: Any, datemode: int) -> dict[tuple[int, int], Any]:
    """构建 .xls 合并单元格填充映射：xlrd merged_cells 为半开区间 (row_lo, row_hi, col_lo, col_hi)，值在 (row_lo, col_lo)。"""
    fill_map: dict[tuple[int, int], Any] = {}
    for row_lo, row_hi, col_lo, col_hi in getattr(sheet, "merged_cells", []):
        top_value = _normalize_xls_value(
            sheet.cell_value(row_lo, col_lo), sheet.cell_type(row_lo, col_lo), datemode
        )
        for r in range(row_lo, row_hi):
            for c in range(col_lo, col_hi):
                if r == row_lo and c == col_lo:
                    continue
                fill_map[(r, c)] = top_value
    return fill_map


def _normalize_xls_value(value: Any, cell_type: int, datemode: int) -> Any:
    """xlrd 单元格值归一：日期类型 / 文本日期串转 ISO，其余原样。"""
    try:
        import xlrd
    except ImportError:
        return value
    if cell_type == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(value, datemode).date().isoformat()
        except Exception:  # XLDateError 等
            return value
    if isinstance(value, str):
        parsed = _parse_human_date(value)
        if parsed is not None:
            return parsed
    return value


# ──────────────────────────────────────────────────────────────
# Word .docx
# ──────────────────────────────────────────────────────────────


def _extract_docx(path: Path) -> tuple[str, str]:
    try:
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError as e:
        raise LlmError("Word 解析依赖缺失，请联系开发者") from e

    try:
        doc = Document(str(path))
    except Exception as e:  # 损坏 / 加密
        raise LlmError(f"无法读取 Word 文件：{e}") from e

    parts = [f"# 文档: {path.stem}\n"]
    # 关键：按 body 原始顺序交错输出段落与表格（docx 单独遍历会丢序）
    for block in _iter_block_items(doc):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if text:
                parts.append(text)
        elif isinstance(block, Table):
            table_md = _table_to_markdown(block)
            if table_md:
                parts.append(table_md)
    return "\n\n".join(parts).strip(), "docx"


def _table_to_markdown(table: Any) -> str:
    """把 Word 表格渲染为 Markdown 表格（首行作表头）；空表返回空串。"""
    rows: list[list[str]] = []
    for row in table.rows:
        rows.append([_cell_str(cell.text) for cell in row.cells])
    rows = [r for r in rows if any(r)]
    if not rows:
        return ""
    header = rows[0]
    col_count = len(header)
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * col_count) + " |"]
    for r in rows[1:]:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def _iter_block_items(doc: Any) -> Iterator[Any]:
    """按 document.element.body 子元素顺序迭代段落 / 表格（python-docx cookbook 写法）。"""
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    parent_elm = doc.element.body
    for child in parent_elm.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, doc)
        elif child.tag == qn("w:tbl"):
            yield Table(child, doc)
